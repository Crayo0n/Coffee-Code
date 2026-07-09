import os

file_path = r"c:\CAFECODE09\Coffee-Code\API\app\routers\estadisticas.py"

content = """
from typing import Optional
from fastapi import Query
from fastapi.responses import FileResponse
import os
import tempfile
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# === RECOVERED AND NEW ENDPOINTS ===

@router.get("/kpis")
def get_kpis(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query_sales = db.query(func.sum(HistorialVenta.total_paid))
    query_orders = db.query(Pedido).join(HistorialVenta, Pedido.id == HistorialVenta.order_id)
    
    if start_date:
        query_sales = query_sales.filter(HistorialVenta.date >= start_date)
        query_orders = query_orders.filter(HistorialVenta.date >= start_date)
    if end_date:
        query_sales = query_sales.filter(HistorialVenta.date <= end_date)
        query_orders = query_orders.filter(HistorialVenta.date <= end_date)
        
    sales_sum = query_sales.scalar()
    ventas_del_dia = Decimal(str(sales_sum)) if sales_sum else Decimal("0.00")
    
    tickets_emitidos = query_orders.count()
    ticket_promedio = (ventas_del_dia / tickets_emitidos) if tickets_emitidos > 0 else Decimal("0.00")
    
    valor_inventario = db.query(func.sum(Producto.price * Producto.stock)).scalar() or Decimal("0.00")
    
    return {
        "ventas_del_dia": ventas_del_dia,
        "tickets_emitidos": tickets_emitidos,
        "ticket_promedio": ticket_promedio,
        "valor_inventario": valor_inventario
    }

@router.get("/historial")
def get_historial(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(HistorialVenta)
    if start_date:
        query = query.filter(HistorialVenta.date >= start_date)
    if end_date:
        query = query.filter(HistorialVenta.date <= end_date)
        
    ventas = query.order_by(HistorialVenta.ticket_id.desc()).all()
    
    historial = []
    for v in ventas:
        pedido = db.query(Pedido).filter(Pedido.id == v.order_id).first()
        cliente = "Público en General"
        mesa = v.table_name
        mesero = pedido.waiter.name if pedido and pedido.waiter else v.cashier_email
        
        historial.append({
            "order_id": v.order_id,
            "date": str(v.date),
            "hour": str(v.hour),
            "cliente": cliente,
            "mesa": mesa,
            "mesero": mesero,
            "articulos": articulos,
            "total": v.total_paid,
            "status": "COMPLETADO"
        })
        
    return {"historial": historial}

@router.get("/pedidos/excel")
def exportar_pedidos_excel(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    query = db.query(Pedido).join(HistorialVenta, Pedido.id == HistorialVenta.order_id)
    if start_date:
        query = query.filter(HistorialVenta.date >= start_date)
    if end_date:
        query = query.filter(HistorialVenta.date <= end_date)
    pedidos_db = query.all()
    
    data = []
    for p in pedidos_db:
        mesa_txt = p.table_name if p.table_name else "Mostrador"
        data.append({
            "ID Pedido": p.id,
            "Mesa": mesa_txt,
            "Estado": p.status,
            "Artículos": len(p.detalles)
        })

    df = pd.DataFrame(data)
    file_path = os.path.join(tempfile.gettempdir(), "reporte_pedidos.xlsx")
    df.to_excel(file_path, index=False, startrow=2)

    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb.active
    ws.title = "Pedidos"
    wb.save(file_path)
    return FileResponse(file_path, filename="reporte_pedidos.xlsx")

@router.get("/pedidos/pdf")
def exportar_pedidos_pdf(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(Pedido).join(HistorialVenta, Pedido.id == HistorialVenta.order_id)
    if start_date:
        query = query.filter(HistorialVenta.date >= start_date)
    if end_date:
        query = query.filter(HistorialVenta.date <= end_date)
    pedidos = query.all()
    
    file_path = os.path.join(tempfile.gettempdir(), "reporte_pedidos.pdf")
    doc = SimpleDocTemplate(file_path, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=50, title="Pedidos", author="Coffee-Code")
    styles = getSampleStyleSheet()
    elementos = []
    elementos.append(Paragraph("<b>Reporte de Pedidos</b>", styles["Normal"]))
    doc.build(elementos)
    return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": "inline; filename=\\"reporte_pedidos.pdf\\""})

@router.get("/productos/excel")
def exportar_productos_excel(categoria: Optional[str] = Query(None), db: Session = Depends(get_db)):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    query = db.query(Producto)
    if categoria and categoria != "ALL":
        query = query.join(Categoria).filter(Categoria.name == categoria)
    productos = query.all()
    
    data = []
    for p in productos:
        data.append({"ID": p.id, "Producto": p.name, "Precio": p.price})

    df = pd.DataFrame(data)
    file_path = os.path.join(tempfile.gettempdir(), "reporte_catalogo.xlsx")
    df.to_excel(file_path, index=False)
    return FileResponse(file_path, filename="reporte_catalogo.xlsx")

@router.get("/productos/pdf")
def exportar_productos_pdf(categoria: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(Producto)
    if categoria and categoria != "ALL":
        query = query.join(Categoria).filter(Categoria.name == categoria)
    productos = query.all()
    
    file_path = os.path.join(tempfile.gettempdir(), "reporte_catalogo.pdf")
    doc = SimpleDocTemplate(file_path, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    doc.build([Paragraph("Catálogo", styles["Normal"])])
    return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": "inline; filename=\\"reporte_catalogo.pdf\\""})

@router.get("/kpis/excel")
def exportar_kpis_excel(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    kpis = get_kpis(start_date, end_date, db)
    data = [{
        "Ventas del Dia": kpis["ventas_del_dia"],
        "Tickets Emitidos": kpis["tickets_emitidos"],
        "Ticket Promedio": kpis["ticket_promedio"],
        "Valor de Inventario": kpis["valor_inventario"]
    }]
    df = pd.DataFrame(data)
    file_path = os.path.join(tempfile.gettempdir(), "reporte_kpis.xlsx")
    df.to_excel(file_path, index=False)
    return FileResponse(file_path, filename="reporte_kpis.xlsx")

@router.get("/kpis/pdf")
def exportar_kpis_pdf(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    kpis = get_kpis(start_date, end_date, db)
    file_path = os.path.join(tempfile.gettempdir(), "reporte_kpis.pdf")
    doc = SimpleDocTemplate(file_path, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    elementos = []
    elementos.append(Paragraph(f"Ventas del Dia: {kpis['ventas_del_dia']}", styles["Normal"]))
    elementos.append(Paragraph(f"Tickets Emitidos: {kpis['tickets_emitidos']}", styles["Normal"]))
    elementos.append(Paragraph(f"Ticket Promedio: {kpis['ticket_promedio']}", styles["Normal"]))
    elementos.append(Paragraph(f"Valor de Inventario: {kpis['valor_inventario']}", styles["Normal"]))
    doc.build(elementos)
    return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": "inline; filename=\\"reporte_kpis.pdf\\""})

"""

with open(file_path, "a", encoding="utf-8") as f:
    f.write(content)

print("Endpoints appended successfully")
