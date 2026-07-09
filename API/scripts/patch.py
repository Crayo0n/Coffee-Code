import re

file_path = r"c:\CAFECODE09\Coffee-Code\API\app\routers\estadisticas.py"

with open(file_path, "r", encoding="utf-8") as f:
    lines = f.readlines()

new_lines = []
skip = False
for i, line in enumerate(lines):
    if "@router.get(\"/pedidos/pdf\")" in line and i > 400: # The second one
        skip = True
        
    if "# === FUNCIONES INTEGRADAS DE ESTADISTICAS ===" in line:
        skip = False
        
    if not skip:
        new_lines.append(line)

content = "".join(new_lines)

content = content.replace(
    """@router.get("/pedidos/excel")
def exportar_pedidos_excel(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    query = db.query(Pedido)
    # Pedido does not have created_at. Filtering by date is not possible at this level.
    pedidos_db = query.all()""",
    """@router.get("/pedidos/excel")
def exportar_pedidos_excel(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    query = db.query(Pedido).join(HistorialVenta, Pedido.id == HistorialVenta.order_id)
    if start_date:
        query = query.filter(HistorialVenta.date >= start_date)
    if end_date:
        query = query.filter(HistorialVenta.date <= end_date)
    pedidos_db = query.all()"""
)

content = content.replace(
    """@router.get("/pedidos/pdf")
def exportar_pedidos_pdf(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(Pedido)
    # Pedido does not have created_at. Filtering by date is not possible at this level.
    pedidos = query.all()""",
    """@router.get("/pedidos/pdf")
def exportar_pedidos_pdf(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(Pedido).join(HistorialVenta, Pedido.id == HistorialVenta.order_id)
    if start_date:
        query = query.filter(HistorialVenta.date >= start_date)
    if end_date:
        query = query.filter(HistorialVenta.date <= end_date)
    pedidos = query.all()"""
)

content = content.replace(
    """@router.get("/productos/excel")
def exportar_productos_excel(db: Session = Depends(get_db)):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    productos = db.query(Producto).all()""",
    """@router.get("/productos/excel")
def exportar_productos_excel(categoria: Optional[str] = Query(None), db: Session = Depends(get_db)):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    query = db.query(Producto)
    if categoria and categoria != "ALL":
        query = query.join(Categoria).filter(Categoria.name == categoria)
    productos = query.all()"""
)

content = content.replace(
    """@router.get("/productos/pdf")
def exportar_productos_pdf(db: Session = Depends(get_db)):
    productos = db.query(Producto).all()""",
    """@router.get("/productos/pdf")
def exportar_productos_pdf(categoria: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(Producto)
    if categoria and categoria != "ALL":
        query = query.join(Categoria).filter(Categoria.name == categoria)
    productos = query.all()"""
)

with open(file_path, "w", encoding="utf-8") as f:
    f.write(content)

print("Patch applied successfully")
