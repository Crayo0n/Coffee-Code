import sys

with open(r'c:\CAFECODE09\Coffee-Code\API\app\routers\estadisticas.py', 'r', encoding='utf-8') as f:
    content = f.read()

historial_endpoints = '''

@router.get("/historial/excel")
def exportar_historial_excel(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(HistorialVenta)
    if start_date: query = query.filter(HistorialVenta.date >= start_date)
    if end_date: query = query.filter(HistorialVenta.date <= end_date)
    ventas = query.order_by(desc(HistorialVenta.date), desc(HistorialVenta.hour)).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial_Ventas"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="90694a", end_color="90694a", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin', color='d3b895'), right=Side(style='thin', color='d3b895'), top=Side(style='thin', color='d3b895'), bottom=Side(style='thin', color='d3b895'))

    headers = ['ID Venta', 'Fecha y Hora', 'Cliente', 'Mesa', 'Mesero', 'Artículos', 'Total', 'Estado']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    data = []
    for v in ventas:
        pedido = db.query(Pedido).filter(Pedido.id == v.order_id).first()
        cliente = "Público en General"
        mesa = v.table_name
        mesero = pedido.waiter.name if pedido and pedido.waiter else v.cashier_email
        estado = pedido.status if pedido else "PAGADO"
        dt_str = f"{v.date.strftime('%Y-%m-%d')} {v.hour.strftime('%H:%M:%S')}"
        data.append([v.ticket_id, dt_str, cliente, mesa, mesero, v.items_count, v.total, estado])

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = center_align
            cell.border = border

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    import tempfile, os
    file_path = os.path.join(tempfile.gettempdir(), "historial_ventas.xlsx")
    wb.save(file_path)
    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=historial_ventas.xlsx"})


@router.get("/historial/pdf")
def exportar_historial_pdf(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(HistorialVenta)
    if start_date: query = query.filter(HistorialVenta.date >= start_date)
    if end_date: query = query.filter(HistorialVenta.date <= end_date)
    ventas = query.order_by(desc(HistorialVenta.date), desc(HistorialVenta.hour)).all()

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    import tempfile, os
    file_path = os.path.join(tempfile.gettempdir(), "historial_ventas.pdf")
    doc = SimpleDocTemplate(file_path, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    elementos = []
    
    title_style = ParagraphStyle(name="TitleStyle", fontSize=18, textColor=colors.HexColor("#90694a"), spaceAfter=20, alignment=1, fontName="Helvetica-Bold")
    elementos.append(Paragraph("Reporte de Historial de Ventas", title_style))
    
    periodo_str = "Periodo: Todos los registros"
    if start_date and end_date: periodo_str = f"Periodo: {start_date} al {end_date}"
    elif start_date: periodo_str = f"Periodo: Desde {start_date}"
    elif end_date: periodo_str = f"Periodo: Hasta {end_date}"
    
    subtitle_style = ParagraphStyle(name="SubTitle", fontSize=12, textColor=colors.HexColor("#555555"), spaceAfter=20, alignment=1)
    elementos.append(Paragraph(periodo_str, subtitle_style))
    
    data = [['ID', 'Fecha y Hora', 'Cliente', 'Mesa', 'Mesero', 'Artículos', 'Total', 'Estado']]
    
    for v in ventas:
        pedido = db.query(Pedido).filter(Pedido.id == v.order_id).first()
        cliente = "Público en General"
        mesa = v.table_name
        mesero = pedido.waiter.name if pedido and pedido.waiter else v.cashier_email
        estado = pedido.status if pedido else "PAGADO"
        dt_str = f"{v.date.strftime('%Y-%m-%d')}\\n{v.hour.strftime('%H:%M:%S')}"
        data.append([str(v.ticket_id), Paragraph(dt_str, styles["Normal"]), cliente, mesa, mesero, str(v.items_count), f"", estado])
    
    t = Table(data, colWidths=[40, 90, 110, 60, 110, 60, 80, 80])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#90694a")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.HexColor("#d3b895")),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#90694a")),
        ('PADDING', (0,0), (-1,-1), 8)
    ]))
    
    for i in range(1, len(data)):
        if i % 2 == 0:
            t.setStyle(TableStyle([('BACKGROUND', (0,i), (-1,i), colors.HexColor("#fcf7fb"))]))
            
    elementos.append(t)
    doc.build(elementos)
    
    return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=historial_ventas.pdf"})
'''

with open(r'c:\CAFECODE09\Coffee-Code\API\app\routers\estadisticas.py', 'a', encoding='utf-8') as f:
    f.write(historial_endpoints)
