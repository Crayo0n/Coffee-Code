import datetime
from decimal import Decimal
from typing import List, Dict, Any
from fastapi import APIRouter, Depends, status
from sqlalchemy import extract, func, case, desc
from sqlalchemy.orm import Session
from app.data.database import get_db
from app.models.historial_venta import HistorialVenta
from app.models.pedido import Pedido
from app.models.producto import Producto
from app.models.categoria import Categoria
from app.models.colaborador import Colaborador
from app.models.detalle_pedido import DetallePedido
from app.data.schemas.dashboard import DashboardMetrics

router = APIRouter(prefix="/api/estadisticas", tags=["Inteligencia de Negocio y Reportes"])

def get_date_range(periodo: str):
    today = datetime.date.today()
    if periodo == "ULTIMOSTRINTA":
        return today - datetime.timedelta(days=30), today
    elif periodo == "SEMANAL":
        return today - datetime.timedelta(days=7), today
    elif periodo == "ANUAL":
        return today.replace(month=1, day=1), today
    else: # ESTEMES (default)
        return today.replace(day=1), today

# --- ENDPOINT PARA KPIs DEL DASHBOARD ---
@router.get("/dashboard-kpis", response_model=DashboardMetrics)
def get_dashboard_kpis(db: Session = Depends(get_db)):
    today = datetime.date.today()
    sales_sum = db.query(func.sum(HistorialVenta.total_paid)).filter(HistorialVenta.date == today).scalar()
    sales_today = Decimal(str(sales_sum)) if sales_sum else Decimal("0.00")
    
    active_orders_count = db.query(Pedido).filter(Pedido.status != "PAGADO").count()
    total_products = db.query(Producto).count()
    total_users = db.query(Colaborador).count()
    
    return {
        "sales_today": sales_today,
        "active_orders_count": active_orders_count,
        "total_products": total_products,
        "total_users": total_users
    }

# --- ENDPOINT PARA REPORTE FINANCIERO (LIBRO AUXILIAR) ---
@router.get("/reporte-financiero")
def get_reporte_financiero(periodo: str = "ESTEMES", db: Session = Depends(get_db)):
    start_date, end_date = get_date_range(periodo)
    sales = db.query(HistorialVenta).filter(
        HistorialVenta.date >= start_date,
        HistorialVenta.date <= end_date
    ).all()
    total_sales = sum(Decimal(str(ticket.total_paid)) for ticket in sales)
    
    semanas = []
    # Calcular semanas reales desde la base de datos basándose en el día del mes
    week1_sales = sum(Decimal(str(t.total_paid)) for t in sales if 1 <= t.date.day <= 7)
    week2_sales = sum(Decimal(str(t.total_paid)) for t in sales if 8 <= t.date.day <= 14)
    week3_sales = sum(Decimal(str(t.total_paid)) for t in sales if 15 <= t.date.day <= 21)
    week4_sales = sum(Decimal(str(t.total_paid)) for t in sales if 22 <= t.date.day <= 31)
    
    ventas_semanales = [
        ("Semana 1", week1_sales, "Inicio de Mes"),
        ("Semana 2", week2_sales, "Estable"),
        ("Semana 3", week3_sales, "Estable"),
        ("Semana 4", week4_sales, "Alta Demanda")
    ]
    for label, ventas, estado in ventas_semanales:
        costo = ventas * Decimal("0.40")
        gasto = Decimal("800.00")
        utilidad = ventas - costo - gasto
        pct = float(utilidad / ventas * 100) if ventas > 0 else 0.0
        semanas.append({
            "periodo": label,
            "ventas_brutas": ventas,
            "costo_insumos": costo,
            "gasto_operativo": gasto,
            "utilidad_neta": utilidad,
            "porcentaje_neto": round(pct, 1),
            "estado": estado
        })
            
    costo_insumos = total_sales * Decimal("0.40")
    gasto_operativo = Decimal("3200.00")
    utilidad_neta = total_sales - costo_insumos - gasto_operativo
    pct_neto = float(utilidad_neta / total_sales * 100) if total_sales > 0 else 0.0
    
    # Histórico de meses anteriores (Abril y Marzo)
    abril_sales = sum(Decimal(str(t.total_paid)) for t in sales if t.date.month == 4 and t.date.year == 2026)
    abril_costo = abril_sales * Decimal("0.40")
    abril_gasto = Decimal("3200.00")
    abril_utilidad = abril_sales - abril_costo - abril_gasto
    abril_pct = float(abril_utilidad / abril_sales * 100) if abril_sales > 0 else 0.0
    
    marzo_sales = sum(Decimal(str(t.total_paid)) for t in sales if t.date.month == 3 and t.date.year == 2026)
    marzo_costo = marzo_sales * Decimal("0.40")
    marzo_gasto = Decimal("3200.00")
    marzo_utilidad = marzo_sales - marzo_costo - marzo_gasto
    marzo_pct = float(marzo_utilidad / marzo_sales * 100) if marzo_sales > 0 else 0.0
    
    historico = [
        {
            "mes": "Abril 2026",
            "ventas_brutas": abril_sales,
            "costo_insumos": abril_costo,
            "gasto_operativo": abril_gasto,
            "utilidad_neta": abril_utilidad,
            "porcentaje_neto": round(abril_pct, 1),
            "estado": "Cerrado"
        },
        {
            "mes": "Marzo 2026",
            "ventas_brutas": marzo_sales,
            "costo_insumos": marzo_costo,
            "gasto_operativo": marzo_gasto,
            "utilidad_neta": marzo_utilidad,
            "porcentaje_neto": round(marzo_pct, 1),
            "estado": "Cerrado"
        }
    ]
    
    return {
        "ventas_brutas": total_sales,
        "costo_insumos": costo_insumos,
        "gasto_operativo": gasto_operativo,
        "utilidad_neta": utilidad_neta,
        "porcentaje_neto": round(pct_neto, 1),
        "semanas": semanas,
        "historico": historico
    }
# --- ENDPOINT PARA REPORTE DE POPULARIDAD DE CATÁLOGO ---
@router.get("/reporte-productos")
def get_reporte_productos(periodo: str = "ESTEMES", db: Session = Depends(get_db)):
    start_date, end_date = get_date_range(periodo)
    results = db.query(
        Producto.name,
        Categoria.name.label("categoria"),
        func.sum(DetallePedido.quantity).label("vendidos"),
        Producto.price
    ).join(
        DetallePedido, Producto.id == DetallePedido.product_id
    ).join(
        Categoria, Producto.category_id == Categoria.id
    ).join(
        Pedido, DetallePedido.order_id == Pedido.id
    ).join(
        HistorialVenta, HistorialVenta.order_id == Pedido.id
    ).filter(
        Pedido.status == "PAGADO",
        HistorialVenta.date >= start_date,
        HistorialVenta.date <= end_date
    ).group_by(
        Producto.id, Categoria.id
    ).order_by(
        func.sum(DetallePedido.quantity).desc()
    ).all()
    
    products_report = []
    for r in results:
        recaudacion = Decimal(str(r.vendidos)) * Decimal(str(r.price))
        costo_promedio = Decimal(str(r.price)) * Decimal("0.40")
        margen_unitario = Decimal(str(r.price)) - costo_promedio
        
        products_report.append({
            "name": r.name,
            "category": r.categoria,
            "units_sold": r.vendidos,
            "price": r.price,
            "revenue": recaudacion,
            "cost_unit": costo_promedio,
            "margin_unit": margen_unitario
        })
        

    return products_report

# --- ENDPOINT PARA FLUJO HORARIO Y MÉTODOS DE PAGO ---
@router.get("/reporte-flujo")
def get_reporte_flujo(periodo: str = "ESTEMES", db: Session = Depends(get_db)):
    start_date, end_date = get_date_range(periodo)
    tarjeta_count = db.query(HistorialVenta).filter(
        HistorialVenta.payment_method == "TARJETA",
        HistorialVenta.date >= start_date,
        HistorialVenta.date <= end_date
    ).count()
    efectivo_count = db.query(HistorialVenta).filter(
        HistorialVenta.payment_method == "EFECTIVO",
        HistorialVenta.date >= start_date,
        HistorialVenta.date <= end_date
    ).count()
    total = tarjeta_count + efectivo_count
    
    tarjeta_pct = (tarjeta_count / total * 100) if total > 0 else 0.0
    efectivo_pct = (efectivo_count / total * 100) if total > 0 else 0.0
    
    sales = db.query(HistorialVenta).filter(
        HistorialVenta.date >= start_date,
        HistorialVenta.date <= end_date
    ).all()
    
    bloques = []
    # Calcular bloques reales desde la base de datos
    b1_sales = [s for s in sales if datetime.time(7, 0) <= s.hour < datetime.time(10, 0)]
    b2_sales = [s for s in sales if datetime.time(10, 0) <= s.hour < datetime.time(13, 0)]
    b3_sales = [s for s in sales if datetime.time(13, 0) <= s.hour < datetime.time(16, 0)]
    b4_sales = [s for s in sales if datetime.time(16, 0) <= s.hour <= datetime.time(23, 59) or datetime.time(0, 0) <= s.hour < datetime.time(7, 0)]
    
    ventas_bloques = [
        ("07:00 AM - 10:00 AM (Apertura y Desayunos)", b1_sales),
        ("10:00 AM - 01:00 PM (Hora Valle Mañana)", b2_sales),
        ("01:00 PM - 04:00 PM (Hora Almuerzo)", b3_sales),
        ("04:00 PM - 07:00 PM (Tarde y Cierre)", b4_sales)
    ]
    for label, block_sales in ventas_bloques:
        pedidos_count = len(block_sales)
        volumen = sum(Decimal(str(s.total_paid)) for s in block_sales)
        
        if pedidos_count >= 10:
            estado = "TRÁFICO CRÍTICO"
        elif pedidos_count >= 5:
            estado = "FLUJO ELEVADO"
        elif pedidos_count >= 2:
            estado = "FLUJO MODERADO"
        else:
            estado = "BAJO FLUJO"
            
        mesa_activa = "N/A"
        if block_sales:
            from collections import Counter
            mesa_counts = Counter(s.table_name for s in block_sales)
            mesa_activa = mesa_counts.most_common(1)[0][0]
            
        bloques.append({
            "periodo": label,
            "pedidos": pedidos_count,
            "volumen_ventas": volumen,
            "tiempo_prep": "N/A",
            "mesa_activa": mesa_activa,
            "estado": estado
        })
            
    # Calcular origen de pedidos
    total_orders = db.query(Pedido).join(HistorialVenta, HistorialVenta.order_id == Pedido.id).filter(
        HistorialVenta.date >= start_date,
        HistorialVenta.date <= end_date
    ).count()
    expo_orders = db.query(Pedido).join(Colaborador, Pedido.waiter_id == Colaborador.id).join(HistorialVenta, HistorialVenta.order_id == Pedido.id).filter(
        Colaborador.role == "MESERO",
        HistorialVenta.date >= start_date,
        HistorialVenta.date <= end_date
    ).count()
    caja_orders = total_orders - expo_orders
    
    expo_pct = round((expo_orders / total_orders * 100), 1) if total_orders > 0 else 0.0
    caja_pct = round((caja_orders / total_orders * 100), 1) if total_orders > 0 else 0.0
    
    return {
        "metodos_pago": {
            "tarjeta_porcentaje": round(tarjeta_pct, 1),
            "efectivo_porcentaje": round(efectivo_pct, 1),
            "tarjeta_tickets": tarjeta_count,
            "efectivo_tickets": efectivo_count
        },
        "origen_pedidos": {
            "expo_pedidos": expo_orders,
            "expo_porcentaje": expo_pct,
            "caja_pedidos": caja_orders,
            "caja_porcentaje": caja_pct
        },
        "bloques": bloques
    }

from typing import Optional
from fastapi import Query
from fastapi.responses import FileResponse
import os
import tempfile
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# === RECOVERED AND NEW ENDPOINTS ===

@router.get("/kpis")
def get_kpis(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query_sales = db.query(func.sum(HistorialVenta.total_paid))
    query_orders = db.query(Pedido).join(HistorialVenta, Pedido.id == HistorialVenta.order_id)
    
    if start_date:
        query_sales = query_sales.filter(HistorialVenta.date >= start_date)
        query_orders = query_orders.filter(HistorialVenta.date >= start_date)
    if end_date:
        query_sales = query_sales.filter(HistorialVenta.date <= end_date)
        query_orders = query_orders.filter(HistorialVenta.date <= end_date)
        
    sales_sum = query_sales.scalar()
    ventas_del_dia = Decimal(str(sales_sum)) if sales_sum else Decimal("0.00")
    
    tickets_emitidos = query_orders.count()
    ticket_promedio = (ventas_del_dia / tickets_emitidos) if tickets_emitidos > 0 else Decimal("0.00")
    
    valor_inventario = db.query(func.sum(Producto.price * Producto.stock)).scalar() or Decimal("0.00")
    
    return {
        "ventas_del_dia": ventas_del_dia,
        "tickets_emitidos": tickets_emitidos,
        "ticket_promedio": ticket_promedio,
        "valor_inventario": valor_inventario
    }

@router.get("/historial")
def get_historial(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(HistorialVenta)
    if start_date:
        query = query.filter(HistorialVenta.date >= start_date)
    if end_date:
        query = query.filter(HistorialVenta.date <= end_date)
        
    ventas = query.order_by(HistorialVenta.ticket_id.desc()).all()
    
    historial = []
    for v in ventas:
        pedido = db.query(Pedido).filter(Pedido.id == v.order_id).first()
        cliente = "Público en General"
        mesa = v.table_name
        mesero = "Caja" # Simplification for now
        if pedido and pedido.waiter_id:
            colab = db.query(Colaborador).filter(Colaborador.id == pedido.waiter_id).first()
            if colab:
                mesero = colab.name
                
        articulos = len(pedido.detalles) if pedido else 0
        
        historial.append({
            "id": v.ticket_id,
            "order_id": v.order_id,
            "fecha_hora": f"{v.date.strftime('%Y-%m-%d')}, {v.hour.strftime('%H:%M:%S')}",
            "cliente": cliente,
            "mesa": mesa,
            "mesero": mesero,
            "articulos": articulos,
            "total": float(v.total_paid),
            "estado": pedido.status if pedido else "Completado"
        })
        
    return {"historial": historial}

@router.get("/pedidos/excel")
def exportar_pedidos_excel(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(Pedido).join(HistorialVenta, Pedido.id == HistorialVenta.order_id)
    if start_date:
        query = query.filter(HistorialVenta.date >= start_date)
    if end_date:
        query = query.filter(HistorialVenta.date <= end_date)
    pedidos = query.all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte_Pedidos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="90694a", end_color="90694a", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin', color='d3b895'), 
                    right=Side(style='thin', color='d3b895'), 
                    top=Side(style='thin', color='d3b895'), 
                    bottom=Side(style='thin', color='d3b895'))

    from openpyxl.drawing.image import Image as ExcelImage
    import os
    logo_path = r"/code/app/assets/logo.png"
    if os.path.exists(logo_path):
        img = ExcelImage(logo_path)
        img.width = 80
        img.height = 80
        ws.add_image(img, 'A1')
    
    ws.row_dimensions[1].height = 65
    ws.merge_cells('C1:D1')
    ws['C1'] = "Reporte de Pedidos"
    ws['C1'].font = Font(size=18, bold=True, color="90694a")
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")

    headers = ['ID Pedido', 'Mesa', 'Estado', 'Artículos']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    
    data = []
    for p in pedidos:
        mesa_txt = p.table_name if p.table_name else "Mostrador"
        data.append([p.id, mesa_txt, p.status, len(p.detalles)])


    for row_num, row_data in enumerate(data, 4):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = center_align
            cell.border = border
            
    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(idx)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column].width = adjusted_width

    file_path = os.path.join(tempfile.gettempdir(), "reporte_pedidos.xlsx")
    wb.save(file_path)
    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=\"reporte_pedidos.xlsx\""})

@router.get("/pedidos/pdf")
def exportar_pedidos_pdf(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    from reportlab.lib.styles import ParagraphStyle
    query = db.query(Pedido).join(HistorialVenta, Pedido.id == HistorialVenta.order_id)
    if start_date:
        query = query.filter(HistorialVenta.date >= start_date)
    if end_date:
        query = query.filter(HistorialVenta.date <= end_date)
    pedidos = query.all()
    
    file_path = os.path.join(tempfile.gettempdir(), "reporte_pedidos.pdf")
    doc = SimpleDocTemplate(file_path, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=50, title="Pedidos", author="Coffee-Code")
    elementos = []
    
    from reportlab.platypus import Image
    logo_path = r"/code/app/assets/logo.png"
    logo = Image(logo_path, width=60, height=60) if __import__('os').path.exists(logo_path) else ""
    title_style = ParagraphStyle(name="TitleStyle", fontSize=18, textColor=colors.HexColor("#90694a"), alignment=1, fontName="Helvetica-Bold")
    title_p = Paragraph(r"Reporte de Pedidos", title_style)
    header_table = Table([[logo, title_p, ""]], colWidths=[80, 400, 80])
    header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    elementos.append(header_table)
    elementos.append(Spacer(1, 15))
    
    data = [["ID Pedido", "Mesa", "Estado", "Artículos"]]
    for p in pedidos:
        mesa_txt = p.table_name if p.table_name else "Mostrador"
        data.append([str(p.id), mesa_txt, str(p.status), str(len(p.detalles))])
        
    t = Table(data, colWidths=[80, 150, 150, 100])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#90694a")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#fcf7fb")),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#d3b895"))
    ]))
    elementos.append(t)
    doc.build(elementos)
    return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": "inline; filename=\"reporte_pedidos.pdf\""})

@router.get("/productos/excel")
def exportar_productos_excel(categoria: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(Producto)
    if categoria and categoria != "ALL":
        query = query.join(Categoria).filter(Categoria.name == categoria)
    productos = query.all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte_Productos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="90694a", end_color="90694a", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin', color='d3b895'), 
                    right=Side(style='thin', color='d3b895'), 
                    top=Side(style='thin', color='d3b895'), 
                    bottom=Side(style='thin', color='d3b895'))

    from openpyxl.drawing.image import Image as ExcelImage
    import os
    logo_path = r"/code/app/assets/logo.png"
    if os.path.exists(logo_path):
        img = ExcelImage(logo_path)
        img.width = 80
        img.height = 80
        ws.add_image(img, 'A1')
    
    ws.row_dimensions[1].height = 65
    ws.merge_cells('C1:D1')
    ws['C1'] = "Reporte de Productos"
    ws['C1'].font = Font(size=18, bold=True, color="90694a")
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")

    headers = ['ID Producto', 'Nombre', 'Categoría', 'Precio']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    
    data = []
    for p in productos:
        cat_name = p.categoria.name if p.categoria else "General"
        data.append([p.id, p.name, cat_name, float(p.price)])


    for row_num, row_data in enumerate(data, 4):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = center_align
            cell.border = border
            
    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(idx)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column].width = adjusted_width

    file_path = os.path.join(tempfile.gettempdir(), "reporte_productos.xlsx")
    wb.save(file_path)
    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=\"reporte_productos.xlsx\""})

@router.get("/productos/pdf")
def exportar_productos_pdf(categoria: Optional[str] = Query(None), db: Session = Depends(get_db)):
    from reportlab.lib.styles import ParagraphStyle
    query = db.query(Producto)
    if categoria and categoria != "ALL":
        query = query.join(Categoria).filter(Categoria.name == categoria)
    productos = query.all()
    
    file_path = os.path.join(tempfile.gettempdir(), "reporte_catalogo.pdf")
    doc = SimpleDocTemplate(file_path, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=50, title="Reporte Coffee-Code", author="Coffee-Code")
    elementos = []
    
    from reportlab.platypus import Image
    logo_path = r"/code/app/assets/logo.png"
    logo = Image(logo_path, width=60, height=60) if __import__('os').path.exists(logo_path) else ""
    title_style = ParagraphStyle(name="TitleStyle", fontSize=18, textColor=colors.HexColor("#90694a"), alignment=1, fontName="Helvetica-Bold")
    title_p = Paragraph(r"Catálogo de Productos", title_style)
    header_table = Table([[logo, title_p, ""]], colWidths=[80, 400, 80])
    header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    elementos.append(header_table)
    elementos.append(Spacer(1, 15))
    
    data = [["ID", "Producto", "Categoría", "Precio"]]
    for p in productos:
        cat_name = p.categoria.name if p.categoria else "General"
        data.append([str(p.id), p.name, cat_name, f"${p.price:.2f}"])
        
    t = Table(data, colWidths=[50, 200, 150, 100])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#90694a")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#fcf7fb")),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#d3b895"))
    ]))
    elementos.append(t)
    doc.build(elementos)
    return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": "inline; filename=\"reporte_catalogo.pdf\""})

@router.get("/kpis/excel")
def exportar_kpis_excel(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    kpis = get_kpis(start_date, end_date, db)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte_KPIs"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="90694a", end_color="90694a", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin', color='d3b895'), 
                    right=Side(style='thin', color='d3b895'), 
                    top=Side(style='thin', color='d3b895'), 
                    bottom=Side(style='thin', color='d3b895'))

    from openpyxl.drawing.image import Image as ExcelImage
    import os
    logo_path = r"/code/app/assets/logo.png"
    if os.path.exists(logo_path):
        img = ExcelImage(logo_path)
        img.width = 80
        img.height = 80
        ws.add_image(img, 'A1')
    
    ws.row_dimensions[1].height = 65
    ws.merge_cells('B1:C1')
    ws['B1'] = "Reporte de KPIs"
    ws['B1'].font = Font(size=18, bold=True, color="90694a")
    ws['B1'].alignment = Alignment(horizontal="center", vertical="center")

    headers = ['Métrica', 'Valor']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    
    data = [
        ["Ventas del Día", kpis['ventas_del_dia']],
        ["Tickets Emitidos", kpis['tickets_emitidos']],
        ["Ticket Promedio", kpis['ticket_promedio']],
        ["Valor de Inventario", kpis['valor_inventario']]
    ]


    for row_num, row_data in enumerate(data, 4):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = center_align
            cell.border = border
            
    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(idx)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column].width = adjusted_width

    file_path = os.path.join(tempfile.gettempdir(), "reporte_kpis.xlsx")
    wb.save(file_path)
    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=\"reporte_kpis.xlsx\""})

@router.get("/kpis/pdf")
def exportar_kpis_pdf(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    from reportlab.lib.styles import ParagraphStyle
    kpis = get_kpis(start_date, end_date, db)
    file_path = os.path.join(tempfile.gettempdir(), "reporte_kpis.pdf")
    doc = SimpleDocTemplate(file_path, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=50, title="Reporte Coffee-Code", author="Coffee-Code")
    elementos = []
    from reportlab.platypus import Image
    logo_path = r"/code/app/assets/logo.png"
    logo = Image(logo_path, width=60, height=60) if __import__('os').path.exists(logo_path) else ""
    title_style = ParagraphStyle(name="TitleStyle", fontSize=18, textColor=colors.HexColor("#90694a"), alignment=1, fontName="Helvetica-Bold")
    title_p = Paragraph(r"Reporte de KPIs", title_style)
    header_table = Table([[logo, title_p, ""]], colWidths=[80, 400, 80])
    header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    elementos.append(header_table)
    elementos.append(Spacer(1, 15))
    
    data = [
        ["Métrica", "Valor"],
        ["Ventas del Día", f"${kpis['ventas_del_dia']:.2f}"],
        ["Tickets Emitidos", str(kpis['tickets_emitidos'])],
        ["Ticket Promedio", f"${kpis['ticket_promedio']:.2f}"],
        ["Valor de Inventario", f"${kpis['valor_inventario']:.2f}"]
    ]
    
    t = Table(data, colWidths=[200, 200])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#90694a")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#fcf7fb")),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#d3b895"))
    ]))
    elementos.append(t)
    doc.build(elementos)
    return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": "inline; filename=\"reporte_kpis.pdf\""})



@router.get("/pedidos/{pedido_id}/ticket/pdf")
def exportar_ticket_pdf(pedido_id: int, db: Session = Depends(get_db)):
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib import colors
    
    pedido = db.query(Pedido).filter(Pedido.id == pedido_id).first()
    if not pedido: return {"error": "Pedido no encontrado"}
    
    venta = db.query(HistorialVenta).filter(HistorialVenta.order_id == pedido_id).first()
    detalles = db.query(DetallePedido).filter(DetallePedido.order_id == pedido_id).all()
    productos = {p.id: p for p in db.query(Producto).all()}
    
    import tempfile, os
    from reportlab.lib.units import mm
    file_path = os.path.join(tempfile.gettempdir(), f"ticket_{pedido_id}.pdf")
    
    # 80mm width for thermal printer
    doc = SimpleDocTemplate(file_path, pagesize=(80*mm, 200*mm), rightMargin=3*mm, leftMargin=3*mm, topMargin=3*mm, bottomMargin=3*mm, title=f"Ticket Pedido #{pedido_id}", author="Coffee-Code")
    elementos = []
    
    styles = getSampleStyleSheet()
    center_style = ParagraphStyle(name="Center", alignment=1)
    title_style = ParagraphStyle(name="Title", alignment=1, fontSize=16, textColor=colors.HexColor("#90694a"), fontName="Helvetica-Bold", spaceAfter=5)
    
    # Logo
    logo_path = r"/code/app/assets/logo.png"
    if os.path.exists(logo_path):
        img = Image(logo_path, width=80, height=80)
        img.hAlign = 'CENTER'
        elementos.append(img)
        elementos.append(Spacer(1, 10))
        
    fecha_str = f"{venta.date} {str(venta.hour)[:11]}" if venta else "N/A"
    estado_str = pedido.status.upper() if pedido.status else "COMPLETADO"
    mesa_str = venta.table_name if venta else (pedido.table_name or "Mostrador")
    mesero_str = pedido.waiter.name if pedido.waiter else "Caja"
    
    # Header text
    elementos.append(Paragraph("Coffee-Code", title_style))
    elementos.append(Paragraph("Ticket de Venta", center_style))
    elementos.append(Paragraph(f"<font size=10>ID: {pedido_id} | {fecha_str}</font>", center_style))
    elementos.append(Spacer(1, 15))
    
    # Client Info
    info_style = ParagraphStyle(name="Info", fontSize=11, leading=14)
    elementos.append(Paragraph(f"<b>Cliente:</b> Público en General", info_style))
    elementos.append(Paragraph(f"<b>Mesa:</b> {mesa_str}", info_style))
    elementos.append(Paragraph(f"<b>Mesero:</b> {mesero_str}", info_style))
    elementos.append(Paragraph(f"<b>Estado:</b> {estado_str}", info_style))
    elementos.append(Spacer(1, 15))
    
    # Table headers
    tabla_data = [["Cant", "Art", "Total"]]
    total = 0
    
    for d in detalles:
        p = productos.get(d.product_id)
        if p:
            sub = float(p.price) * int(d.quantity)
            total += sub
            tabla_data.append([str(d.quantity), p.name, f"${sub:.2f}"])
            
    # Items Table
    t = Table(tabla_data, colWidths=[25, 125, 50])
    t.setStyle(TableStyle([
        ('LINEBELOW', (0,0), (-1,0), 1, colors.HexColor("#dddddd")),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (2,0), (2,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 6)
    ]))
    elementos.append(t)
    elementos.append(Spacer(1, 10))
    
    # Totals
    total_data = []
    if venta and venta.tips:
        total_data.append(["Propina:", f"${float(venta.tips):.2f}"])
        total += float(venta.tips)
    total_data.append(["TOTAL:", f"${total:.2f}"])
    
    total_table = Table(total_data, colWidths=[65, 65])
    total_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,-1), (-1,-1), 12),
        ('LINEABOVE', (0,-1), (-1,-1), 1, colors.HexColor("#352728")),
        ('TOPPADDING', (0,-1), (-1,-1), 8)
    ]))
    
    layout = Table([["", total_table]], colWidths=[70, 130])
    elementos.append(layout)
    
    elementos.append(Spacer(1, 30))
    elementos.append(Paragraph("<font size=10 color='#888888'>¡Gracias por su compra!</font>", center_style))
    
    doc.build(elementos)
    return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": f"inline; filename=\"ticket_{pedido_id}.pdf\""})


@router.get("/historial/excel")
def exportar_historial_excel(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(HistorialVenta)
    if start_date: query = query.filter(HistorialVenta.date >= start_date)
    if end_date: query = query.filter(HistorialVenta.date <= end_date)
    ventas = query.order_by(desc(HistorialVenta.date), desc(HistorialVenta.hour)).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial_Ventas"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="90694a", end_color="90694a", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin', color='d3b895'), right=Side(style='thin', color='d3b895'), top=Side(style='thin', color='d3b895'), bottom=Side(style='thin', color='d3b895'))

    from openpyxl.drawing.image import Image as ExcelImage
    import os
    logo_path = r"/code/app/assets/logo.png"
    if os.path.exists(logo_path):
        img = ExcelImage(logo_path)
        img.width = 80
        img.height = 80
        ws.add_image(img, 'A1')
    
    ws.row_dimensions[1].height = 65
    ws.merge_cells('C1:F1')
    ws['C1'] = "Reporte de Historial de Ventas"
    ws['C1'].font = Font(size=18, bold=True, color="90694a")
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")

    headers = ['ID Venta', 'Fecha y Hora', 'Cliente', 'Mesa', 'Mesero', 'Artículos', 'Total', 'Estado']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    data = []
    for v in ventas:
        pedido = db.query(Pedido).filter(Pedido.id == v.order_id).first()
        cliente = "Público en General"
        mesa = v.table_name
        mesero = pedido.waiter.name if pedido and pedido.waiter else v.cashier_email
        estado = pedido.status if pedido else "PAGADO"
        dt_str = f"{v.date.strftime('%Y-%m-%d')} {v.hour.strftime('%H:%M:%S')}"
        data.append([v.ticket_id, dt_str, cliente, mesa, mesero, len(pedido.detalles) if pedido else 0, f"${v.total_paid:.2f}", estado])

    for row_num, row_data in enumerate(data, 4):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = center_align
            cell.border = border

    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(idx)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    import tempfile, os
    file_path = os.path.join(tempfile.gettempdir(), "historial_ventas.xlsx")
    wb.save(file_path)
    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=historial_ventas.xlsx"})


@router.get("/historial/pdf")
def exportar_historial_pdf(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None), db: Session = Depends(get_db)):
    query = db.query(HistorialVenta)
    if start_date: query = query.filter(HistorialVenta.date >= start_date)
    if end_date: query = query.filter(HistorialVenta.date <= end_date)
    ventas = query.order_by(desc(HistorialVenta.date), desc(HistorialVenta.hour)).all()

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    import tempfile, os
    file_path = os.path.join(tempfile.gettempdir(), "historial_ventas.pdf")
    doc = SimpleDocTemplate(file_path, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    elementos = []
    
    
    from reportlab.platypus import Image
    logo_path = r"/code/app/assets/logo.png"
    logo = Image(logo_path, width=60, height=60) if __import__('os').path.exists(logo_path) else ""
    title_style = ParagraphStyle(name="TitleStyle", fontSize=18, textColor=colors.HexColor("#90694a"), alignment=1, fontName="Helvetica-Bold")
    title_p = Paragraph(r"Reporte de Historial de Ventas", title_style)
    header_table = Table([[logo, title_p, ""]], colWidths=[80, 540, 80])
    header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    elementos.append(header_table)
    elementos.append(Spacer(1, 15))

    
    periodo_str = "Periodo: Todos los registros"
    if start_date and end_date: periodo_str = f"Periodo: {start_date} al {end_date}"
    elif start_date: periodo_str = f"Periodo: Desde {start_date}"
    elif end_date: periodo_str = f"Periodo: Hasta {end_date}"
    
    subtitle_style = ParagraphStyle(name="SubTitle", fontSize=12, textColor=colors.HexColor("#555555"), spaceAfter=20, alignment=1)
    elementos.append(Paragraph(periodo_str, subtitle_style))
    
    data = [['ID', 'Fecha y Hora', 'Cliente', 'Mesa', 'Mesero', 'Artículos', 'Total', 'Estado']]
    
    for v in ventas:
        pedido = db.query(Pedido).filter(Pedido.id == v.order_id).first()
        cliente = "Público en General"
        mesa = v.table_name
        mesero = pedido.waiter.name if pedido and pedido.waiter else v.cashier_email
        estado = pedido.status if pedido else "PAGADO"
        dt_str = f"{v.date.strftime('%Y-%m-%d')}\n{v.hour.strftime('%H:%M:%S')}"
        data.append([str(v.ticket_id), Paragraph(dt_str, styles["Normal"]), cliente, mesa, mesero, str(len(pedido.detalles) if pedido else 0), f"${v.total_paid:.2f}", estado])
    
    t = Table(data, colWidths=[40, 90, 110, 60, 110, 60, 80, 80])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#90694a")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.HexColor("#d3b895")),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#90694a")),
        ('PADDING', (0,0), (-1,-1), 8)
    ]))
    
    for i in range(1, len(data)):
        if i % 2 == 0:
            t.setStyle(TableStyle([('BACKGROUND', (0,i), (-1,i), colors.HexColor("#fcf7fb"))]))
            
    elementos.append(t)
    doc.build(elementos)
    return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": "inline; filename=historial_ventas.pdf"})


@router.get("/pedidos/{pedido_id}/detalles")
def get_pedido_detalles(pedido_id: int, db: Session = Depends(get_db)):
    pedido = db.query(Pedido).filter(Pedido.id == pedido_id).first()
    if not pedido: return {"error": "Pedido no encontrado"}
    
    venta = db.query(HistorialVenta).filter(HistorialVenta.order_id == pedido_id).first()
    detalles = db.query(DetallePedido).filter(DetallePedido.order_id == pedido_id).all()
    productos = {p.id: p for p in db.query(Producto).all()}
    
    items = []
    for d in detalles:
        p = productos.get(d.product_id)
        if p:
            items.append({
                "nombre": p.name,
                "cantidad": d.quantity,
                "precio": float(p.price),
                "importe": float(p.price) * int(d.quantity)
            })
            
    cliente = "Público en General"
    mesa = venta.table_name if venta else (pedido.table_name if pedido.table_name else "Mostrador")
    mesero = pedido.waiter.name if pedido.waiter else "Caja"
    
    return {
        "id_venta": venta.ticket_id if venta else pedido.id,
        "fecha": str(venta.date) if venta else "",
        "hora": str(venta.hour) if venta else "",
        "cliente": cliente,
        "mesa": mesa,
        "mesero": mesero,
        "items": items,
        "total": venta.total_paid if venta else sum(i["importe"] for i in items),
        "estado": pedido.status
    }

@router.get("/graficas/excel")
def exportar_graficas_excel(periodo: str = "ESTEMES", db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile
    import os
    
    fin_data = get_reporte_financiero(periodo, db)
    prod_data = get_reporte_productos(periodo, db)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte_Financiero"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="90694a", end_color="90694a", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin', color='d3b895'), 
                    right=Side(style='thin', color='d3b895'), 
                    top=Side(style='thin', color='d3b895'), 
                    bottom=Side(style='thin', color='d3b895'))
                    
    ws.append(["Periodo", "Ingresos", "Costos Insumos", "Gasto Operativo", "Utilidad Neta"])
    for col_num in range(1, 6):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        
    for idx, s in enumerate(fin_data['semanas'], 2):
        ws.cell(row=idx, column=1, value=s['periodo'])
        ws.cell(row=idx, column=2, value=float(s['ventas_brutas']))
        ws.cell(row=idx, column=3, value=float(s['costo_insumos']))
        ws.cell(row=idx, column=4, value=float(s['gasto_operativo']))
        ws.cell(row=idx, column=5, value=float(s['utilidad_neta']))
        
    ws2 = wb.create_sheet(title="Top_Productos")
    ws2.append(["Producto", "Categoría", "Unidades Vendidas", "Recaudación"])
    for col_num in range(1, 5):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        
    for idx, p in enumerate(prod_data[:10], 2):
        ws2.cell(row=idx, column=1, value=p['name'])
        ws2.cell(row=idx, column=2, value=p['category'])
        ws2.cell(row=idx, column=3, value=p['units_sold'])
        ws2.cell(row=idx, column=4, value=float(p['revenue']))
        
    file_path = os.path.join(tempfile.gettempdir(), "reporte_graficas.xlsx")
    wb.save(file_path)
    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=\"reporte_graficas.xlsx\""})


@router.get("/graficas/pdf")
def exportar_graficas_pdf(periodo: str = "ESTEMES", db: Session = Depends(get_db)):
    from reportlab.lib.styles import ParagraphStyle
    import tempfile
    import os
    
    fin_data = get_reporte_financiero(periodo, db)
    prod_data = get_reporte_productos(periodo, db)
    
    file_path = os.path.join(tempfile.gettempdir(), "reporte_graficas.pdf")
    doc = SimpleDocTemplate(file_path, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=50, title="Reporte Gráficas", author="Coffee-Code")
    elementos = []
    
    title_style = ParagraphStyle(name="TitleStyle", fontSize=16, textColor=colors.HexColor("#90694a"), alignment=1, fontName="Helvetica-Bold")
    elementos.append(Paragraph(r"Resumen Financiero Mensual", title_style))
    elementos.append(Spacer(1, 15))
    
    data_fin = [["Periodo", "Ingresos", "Costos Insumos", "Gasto Operativo", "Utilidad Neta"]]
    for s in fin_data['semanas']:
        data_fin.append([s['periodo'], f"${s['ventas_brutas']:.2f}", f"${s['costo_insumos']:.2f}", f"${s['gasto_operativo']:.2f}", f"${s['utilidad_neta']:.2f}"])
        
    t_fin = Table(data_fin, colWidths=[100, 100, 100, 100, 100])
    t_fin.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#90694a")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#d3b895"))
    ]))
    elementos.append(t_fin)
    elementos.append(Spacer(1, 30))
    
    elementos.append(Paragraph(r"Top 10 Productos Más Vendidos", title_style))
    elementos.append(Spacer(1, 15))
    
    data_prod = [["Producto", "Categoría", "Unidades Vendidas", "Recaudación"]]
    for p in prod_data[:10]:
        data_prod.append([p['name'], p['category'], str(p['units_sold']), f"${p['revenue']:.2f}"])
        
    t_prod = Table(data_prod, colWidths=[150, 100, 100, 100])
    t_prod.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#90694a")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#d3b895"))
    ]))
    elementos.append(t_prod)
    
    doc.build(elementos)
    return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": "inline; filename=\"reporte_graficas.pdf\""})
